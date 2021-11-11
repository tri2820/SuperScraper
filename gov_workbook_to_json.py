from dataclasses import dataclass, field
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from pandas.core.frame import DataFrame
from functools import lru_cache
import logging
logging.basicConfig(level=logging.DEBUG)
import os

@dataclass
class GovWorkbookSpecs:    
    fileName: str
    sheetName: str
    # Pandas index
    valueStartRow: int
    droppedRows: list[int] = field(default_factory=list)

@lru_cache
def getWorksheet(fileName, sheetName):
    logging.info(f'Loading {fileName}, {sheetName} (lite)')
    return pd.read_excel(fileName, sheetName, header=None)

@lru_cache
def loadWorkbook(fileName):
    logging.info(f'Loading {fileName}')
    return load_workbook(fileName)

def readInfo(spec : GovWorkbookSpecs):
    workbook = loadWorkbook(spec.fileName)
    worksheet = workbook[spec.sheetName]
    mergedBounds = [(a-1,b-1,c-1,d-1) for (a,b,c,d) in map(lambda r: r.bounds, worksheet.merged_cells.ranges)]
    groupedCols = getGroupedCols(worksheet)
    return mergedBounds, groupedCols

def buildPrefixes(spec : GovWorkbookSpecs, worksheet : DataFrame) -> list[tuple[str]]:
    keyArea = worksheet.iloc[0:spec.valueStartRow]
    prefixes = []
    for col in keyArea: prefixes.append(tuple(keyArea[col].dropna().astype(str)))
    return prefixes

def fill(worksheet, mergedBound):
    col_min, row_min, col_max, row_max = mergedBound
    forwardedValue = worksheet.iloc[row_min][col_min]
    worksheet.loc[row_min:row_max, col_min:col_max] = forwardedValue

@dataclass(eq=True, order=True, frozen=True)
class GroupedCol:
    min: int
    max: int
    owner: int = None

def merge(outlines: set[GroupedCol]):
    outlines = sorted(list(outlines))
    result : list[GroupedCol] = []
    for o in outlines:
        mismatched = result == [] or result[-1].max+1 < o.min
        if mismatched: result.append(o)
        else: result[-1] = GroupedCol(result[-1].min, o.max)
    return result

def getGroupedCols(ws: Worksheet):
    dim = ws.column_dimensions
    # TODO: take care of groups of groups of grouped cols
    maxlevel = max([d.outlineLevel for d in dim.values()])
    bounds = {GroupedCol(d.min, d.max) for d in dim.values() if d.outlineLevel == 1}
    if maxlevel == 2:
        bounds2 = {GroupedCol(d.min, d.max) for d in dim.values() if d.outlineLevel == 2}
        for b in bounds2: bounds.discard(GroupedCol(b.max+1, b.max+1))
        bounds |= bounds2
    groups = merge(bounds)
    groups = [GroupedCol(g.min-1, g.max-1) for g in groups]
    return groups

def commonHead(tuples):
    if () in tuples or not tuples: return (), tuples
    heads = [t[0] for t in tuples]
    tails = [t[1:] for t in tuples]
    e = heads[0]
    same = heads.count(e) == len(heads)
    if not same: return (), tuples
    
    common, rests = commonHead(tails)
    return (e,) + common, rests

def addGroupedColsToPrefixes(prefixes, groupedCols : list[GroupedCol]):
    result = []
    _prefixes = prefixes

    dups = getDups(_prefixes)
    prefixes = [s.strip() for p in _prefixes for s in p ]
    for id, group in enumerate(groupedCols):
        rangedPrefixes = _prefixes[group.min:group.max+1]
        

        dupsInRange = getDups(rangedPrefixes)
        rangedPrefixes = [p+(f'#{i}',) if p in dupsInRange else p for (i,p) in enumerate(rangedPrefixes)]

        common, rests = commonHead(rangedPrefixes)
        distinctable = set(rangedPrefixes).intersection(dups) == set()
        withoutGroupID = distinctable
        keys = rangedPrefixes if withoutGroupID else [common+(f'G{id}',)+r for r in rests]

        result += keys

    return result

def allGroupedCols(groupedCols : list, i = 0):
    if not groupedCols: return []
    g = groupedCols[0]
    return [GroupedCol(i,i) for i in range(i, g.min)] + [g] + allGroupedCols(groupedCols[1:], g.max+1)

def toWorksheet(spec : GovWorkbookSpecs):
    logging.info(f'Converting {spec.fileName}/{spec.sheetName}')
    _spec = spec
    mergedBounds, groupedCols = readInfo(_spec)
    worksheet : DataFrame = getWorksheet(_spec.fileName, _spec.sheetName)    
    worksheet.replace(r'\n|\\n',' ', regex=True, inplace=True)
    mergedBounds = [(col_min, row_min, col_max, row_max) for (col_min, row_min, col_max, row_max) in mergedBounds if col_max < worksheet.shape[1]]
    for bound in mergedBounds: fill(worksheet, bound)
    worksheet.drop(spec.droppedRows, inplace=True)
    _spec.valueStartRow -= len(_spec.droppedRows)
    prefixes = buildPrefixes(_spec, worksheet)
    groupedCols.append(GroupedCol(worksheet.shape[1], worksheet.shape[1]))
    groupedCols = allGroupedCols(groupedCols)
    prefixes = addGroupedColsToPrefixes(prefixes, groupedCols)
    worksheet = worksheet[_spec.valueStartRow:]
    worksheet.columns = ['->'.join(p) for p in prefixes]
    worksheet = worksheet.dropna(axis=1, how='all')
    return worksheet

def write(fileName, content):
    os.makedirs(os.path.dirname(fileName), exist_ok=True)
    logging.info(f'Writing to {fileName} {hash(content)}')
    with open(fileName,'w') as f:
            f.write(json)
            f.close()

def getDups(items):
    seen = set()
    dups = set()
    for it in items:
        store = seen.add if not it in seen else dups.add
        store(it)
    return dups


if __name__ == '__main__':
    specs = [
        GovWorkbookSpecs('workbook/Annual fund-level superannuation statistics June 2020.xlsx', 'Table 1', valueStartRow=7, droppedRows=[0,1,2,3,5,6]),
        GovWorkbookSpecs('workbook/Annual fund-level superannuation statistics June 2020.xlsx', 'Table 2', valueStartRow=7, droppedRows=[0,1,3,5,6]),
        GovWorkbookSpecs('workbook/Annual fund-level superannuation statistics June 2020.xlsx', 'Table 3', valueStartRow=7, droppedRows=[0,1,3,5,6]),
        GovWorkbookSpecs('workbook/Annual fund-level superannuation statistics June 2020.xlsx', 'Table 4', valueStartRow=7, droppedRows=[0,1,3,5,6]),
        GovWorkbookSpecs('workbook/Annual fund-level superannuation statistics June 2020.xlsx', 'Table 5', valueStartRow=7, droppedRows=[0,1,3,5,6]),
        GovWorkbookSpecs('workbook/Annual fund-level superannuation statistics June 2020.xlsx', 'Table 6', valueStartRow=9, droppedRows=[0,1,2,5,7,8]),
        GovWorkbookSpecs('workbook/Annual fund-level superannuation statistics June 2020.xlsx', 'Table 7', valueStartRow=8, droppedRows=[0,1,2,4,6,7]),
        GovWorkbookSpecs('workbook/Annual fund-level superannuation statistics June 2020.xlsx', 'Table 8', valueStartRow=7, droppedRows=[0,1,2,3,5,6]),
        GovWorkbookSpecs('workbook/Annual fund-level superannuation statistics June 2020.xlsx', 'Table 9', valueStartRow=8, droppedRows=[0,1,2,4,6,7]),
        GovWorkbookSpecs('workbook/Annual fund-level superannuation statistics June 2020.xlsx', 'Table 10', valueStartRow=8, droppedRows=[0,1,2,4,6,7]),
        GovWorkbookSpecs('workbook/Annual fund-level superannuation statistics June 2020.xlsx', 'Table 11', valueStartRow=8, droppedRows=[0,1,2,4,6,7]),
        GovWorkbookSpecs('workbook/Annual fund-level superannuation statistics June 2020.xlsx', 'Table 12', valueStartRow=8, droppedRows=[0,1,2,4,6,7]),
        GovWorkbookSpecs('workbook/Annual fund-level superannuation statistics June 2020.xlsx', 'Table 13', valueStartRow=8, droppedRows=[0,1,2,4,6,7]),

        GovWorkbookSpecs('workbook/List of RSES and RSE Licensees and MySuper Authorised products and ERFs 23 August 2021.xlsx', 'List of RSE', valueStartRow=1),
        GovWorkbookSpecs('workbook/List of RSES and RSE Licensees and MySuper Authorised products and ERFs 23 August 2021.xlsx', 'List of Licensee', valueStartRow=1),
        GovWorkbookSpecs('workbook/List of RSES and RSE Licensees and MySuper Authorised products and ERFs 23 August 2021.xlsx', 'MySuper Products', valueStartRow=1),
        GovWorkbookSpecs('workbook/List of RSES and RSE Licensees and MySuper Authorised products and ERFs 23 August 2021.xlsx', 'Eligible Rollover Funds', valueStartRow=1),

        GovWorkbookSpecs('workbook/Quarterly Mysuper statistics September 2019 - March 2021 reissued.xlsx', 'Table 1a', valueStartRow=6, droppedRows=[0,1,4,5]),
        GovWorkbookSpecs('workbook/Quarterly Mysuper statistics September 2019 - March 2021 reissued.xlsx', 'Table 1b', valueStartRow=6, droppedRows=[0,1,4,5]),
        GovWorkbookSpecs('workbook/Quarterly Mysuper statistics September 2019 - March 2021 reissued.xlsx', 'Table 2a', valueStartRow=7, droppedRows=[0,1,3,5,6]),
        GovWorkbookSpecs('workbook/Quarterly Mysuper statistics September 2019 - March 2021 reissued.xlsx', 'Table 2b', valueStartRow=8, droppedRows=[0,1,3,5,6]),
        GovWorkbookSpecs('workbook/Quarterly Mysuper statistics September 2019 - March 2021 reissued.xlsx', 'Table 3', valueStartRow=8, droppedRows=[0,1,3,5,6]),
        GovWorkbookSpecs('workbook/Quarterly Mysuper statistics September 2019 - March 2021 reissued.xlsx', 'Table 4', valueStartRow=5, droppedRows=[0,1,3,4]),
        GovWorkbookSpecs('workbook/Quarterly Mysuper statistics September 2019 - March 2021 reissued.xlsx', 'Table 5', valueStartRow=6, droppedRows=[0,1,2,4,5]),
        GovWorkbookSpecs('workbook/Quarterly Mysuper statistics September 2019 - March 2021 reissued.xlsx', 'Table 6', valueStartRow=6, droppedRows=[0,1,2,4,5]),
        GovWorkbookSpecs('workbook/Quarterly Mysuper statistics September 2019 - March 2021 reissued.xlsx', 'Table 7', valueStartRow=7, droppedRows=[0,1,2,5,6]),
        GovWorkbookSpecs('workbook/Quarterly Mysuper statistics September 2019 - March 2021 reissued.xlsx', 'Table 8', valueStartRow=6, droppedRows=[0,1,2,4,5]),
    ]
    for spec in specs:
        worksheet = toWorksheet(spec)
        
        keys = worksheet.columns
        dups = getDups(keys)
        if dups:
            raise Exception('Cannot convert json records, there are duplications in keys: ', dups, keys)

        json = worksheet.to_json(orient='records')
        path = os.path.join('result', spec.fileName, f'{spec.sheetName}.json')
        write(path, json)